"""
data_loader.py
Modul untuk memuat data dari file Excel lokal atau Google Sheets (published CSV).
Mengembalikan dua DataFrame: df_nilai (persentase) dan df_ranking.
"""

import io
import pandas as pd
import openpyxl
import requests


# ---------------------------------------------------------------------------
# Helper: parse raw rows (list of lists) → (df_nilai, df_ranking)
# ---------------------------------------------------------------------------

def _parse_tables(rows):
    """
    Menerima list-of-lists (setiap inner list = 1 baris data).
    Mendukung dua format:
      - Format lama: kolom ke-1 (index 1) = 'Kabupaten/Kota', tanggal mulai kolom ke-2
      - Format baru: kolom ke-0 (index 0) = 'Kabupaten/Kota', tanggal mulai kolom ke-1
    Return (df_nilai, df_ranking).
    """
    tables = []
    i = 0
    while i < len(rows):
        row = rows[i]
        # Cari header row — cek kolom 0 atau 1
        header_col = None
        if len(row) > 0 and str(row[0]).strip() == "Kabupaten/Kota":
            header_col = 0
        elif len(row) > 1 and str(row[1]).strip() == "Kabupaten/Kota":
            header_col = 1
        
        if header_col is not None:
            data_start_col = header_col + 1
            # Ambil tanggal dari kolom data_start_col dst
            tanggal_raw = [v for v in row[data_start_col:] if v is not None and str(v).strip() != ""]
            tanggal = _parse_tanggal(tanggal_raw)
            n_cols = len(tanggal)

            data_rows = []
            i += 1
            while i < len(rows):
                r = rows[i]
                nama = r[header_col] if len(r) > header_col else None
                if nama is None or str(nama).strip() == "" or str(nama).strip() == "Kabupaten/Kota":
                    break
                vals = []
                for c in range(data_start_col, data_start_col + n_cols):
                    v = r[c] if c < len(r) else None
                    vals.append(_to_float(v))
                data_rows.append((str(nama).strip(), vals))
                i += 1

            if data_rows:
                idx = [r[0] for r in data_rows]
                data = [r[1] for r in data_rows]
                df = pd.DataFrame(data, index=idx, columns=tanggal)
                tables.append(df)
        else:
            i += 1

    if len(tables) >= 2:
        df_nilai = tables[0]
        df_ranking = tables[1]
    elif len(tables) == 1:
        df_nilai = tables[0]
        df_ranking = df_nilai.rank(axis=0, ascending=False, method="min").astype(int)
    else:
        raise ValueError("Tidak ditemukan tabel dengan header 'Kabupaten/Kota'.")

    return df_nilai, df_ranking


def _parse_tanggal(tanggal_raw):
    """Konversi list angka tanggal mentah ke label 'DD Mon', mendeteksi rollover bulan otomatis."""
    tanggal = []
    prev_num = None
    current_month = None
    for t in tanggal_raw:
        if _is_number(t):
            num = int(float(t))
            if current_month is None:
                # Angka pertama >= 18 = Juni, < 18 = Juli
                current_month = "Jun" if num >= 18 else "Jul"
            elif prev_num is not None and num < prev_num - 5:
                # Deteksi rollover: angka turun tajam → bulan baru
                if current_month == "Jun":
                    current_month = "Jul"
                elif current_month == "Jul":
                    current_month = "Agu"
            prev_num = num
            tanggal.append(f"{num:02d} {current_month}")
        else:
            tanggal.append(str(t))
    return tanggal




def _is_number(v):
    try:
        float(v)
        return True
    except (ValueError, TypeError):
        return False


def _to_float(v):
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


# ---------------------------------------------------------------------------
# Loader: Excel file
# ---------------------------------------------------------------------------

def load_from_excel(file_path_or_buffer):
    """
    Memuat data dari file Excel (.xlsx).
    file_path_or_buffer: path string atau BytesIO (dari st.file_uploader).
    """
    wb = openpyxl.load_workbook(file_path_or_buffer, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    for r in range(1, ws.max_row + 1):
        row = []
        for c in range(1, ws.max_column + 1):
            row.append(ws.cell(r, c).value)
        rows.append(row)

    return _parse_tables(rows)


# ---------------------------------------------------------------------------
# Loader: Google Sheets (published CSV)
# ---------------------------------------------------------------------------

def load_from_gsheet(url):
    """
    Memuat data dari Google Sheets URL.
    Secara otomatis mengkonversi URL edit → URL export CSV.
    """
    csv_url = _to_csv_export_url(url)

    response = requests.get(csv_url, timeout=30)
    response.raise_for_status()

    # Parse CSV → list of lists
    import csv
    reader = csv.reader(io.StringIO(response.text))
    rows = [row for row in reader]

    # Pad rows agar panjangnya sama
    max_len = max(len(r) for r in rows) if rows else 0
    rows = [r + [None] * (max_len - len(r)) for r in rows]

    return _parse_tables(rows)


def _to_csv_export_url(url):
    """Konversi berbagai format Google Sheets URL → export CSV URL."""
    import re

    # Ekstrak spreadsheet ID
    match = re.search(r"/d/([a-zA-Z0-9_-]+)", url)
    if not match:
        raise ValueError(f"URL Google Sheets tidak valid: {url}")
    sheet_id = match.group(1)

    # Ekstrak gid (jika ada)
    gid_match = re.search(r"gid=(\d+)", url)
    gid = gid_match.group(1) if gid_match else "0"

    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"


# ---------------------------------------------------------------------------
# Utility: Pisahkan baris PROVINSI RIAU dari data kabupaten
# ---------------------------------------------------------------------------

def split_provinsi(df):
    """
    Pisahkan baris 'PROVINSI RIAU' dari DataFrame.
    Return (df_kabupaten, series_provinsi_or_None).
    """
    prov_label = "PROVINSI RIAU"
    if prov_label in df.index:
        sr = df.loc[prov_label]
        df_kab = df.drop(prov_label)
        return df_kab, sr
    return df, None
